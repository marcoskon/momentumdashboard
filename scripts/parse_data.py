"""
parse_data.py
=============
Convierte los archivos xlsx/xls de la carpeta raw/ en JSON consumibles
por el dashboard.  Se ejecuta automáticamente vía GitHub Actions.

Estructura esperada en raw/:
  raw/
    GESTION_MOMENTUM_20260507.xlsx
    GESTION_MOMENTUM_20260508.xlsx
    Balance_07-05-2026.xls
    Balance_08-05-2026.xls

Por cada par de archivos del mismo día genera:
  data/2026-05-07.json

Y actualiza:
  data/index.json  ← lista de fechas disponibles
"""

import json
import os
import re
from pathlib import Path
from datetime import datetime

import openpyxl
import xlrd

# ─────────────────────────────────────────────
# RUTAS
# ─────────────────────────────────────────────
ROOT     = Path(__file__).parent.parent
RAW_DIR  = ROOT / "raw"
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fmt_date_iso(raw: str) -> str | None:
    """Convierte '20260507' o '07-05-2026' a '2026-05-07'."""
    raw = raw.strip()
    if re.match(r'^\d{8}$', raw):                        # 20260507
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', raw)     # 07-05-2026
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None

def date_from_filename(fname: str) -> str | None:
    """Extrae fecha ISO de cualquier nombre de archivo razonable."""
    # GESTION_MOMENTUM_20260507  →  20260507
    m = re.search(r'(\d{8})', fname)
    if m:
        return fmt_date_iso(m.group(1))
    # Balance_07-05-2026         →  07-05-2026
    m = re.search(r'(\d{2}-\d{2}-\d{4})', fname)
    if m:
        return fmt_date_iso(m.group(1))
    return None

def cell(ws, row: int, col: int):
    """Lee un valor de celda de forma segura (openpyxl)."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v

def safe_float(v) -> float | None:
    try:
        return float(str(v).replace(',', '.').replace('%', ''))
    except Exception:
        return None

def fmt_m(v, decimals=3) -> str:
    """Formatea número en billones/millones para display."""
    if v is None:
        return "—"
    v = float(v)
    if abs(v) >= 1e9:
        return f"${v/1e9:.{decimals}f} B"
    if abs(v) >= 1e6:
        return f"${v/1e6:.1f} M"
    return f"${v:,.0f}"

def pct_delta(hoy, ayer) -> str:
    if not hoy or not ayer or ayer == 0:
        return "—"
    d = (hoy - ayer) / abs(ayer) * 100
    sign = "▲ +" if d >= 0 else "▼ "
    return f"{sign}{d:.2f}%"

# ─────────────────────────────────────────────────────────────────────────────
# PARSERS
# Ajustá las referencias de celdas (row, col) según la estructura real de tus
# archivos.  Usá Excel para verificar: F5 → escribí la celda, p.ej. B5.
# ─────────────────────────────────────────────────────────────────────────────

def parse_gestion(path: Path) -> dict:
    """
    Lee el archivo GESTION_MOMENTUM_YYYYMMDD.xlsx.

    ┌─────────────────────────────────────────────────────────────┐
    │  CONFIGURACIÓN — ajustá las referencias de celdas aquí     │
    └─────────────────────────────────────────────────────────────┘
    Abrí el xlsx en Excel, encontrá cada dato y anotá fila/columna.
    Ejemplo: si el VCP Clase A está en la hoja "VCP" celda C5,
             ponés sheet="VCP", row=5, col=3
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Imprime hojas disponibles para ayudarte a configurar ──
    print(f"  Hojas en {path.name}: {wb.sheetnames}")

    # ── Intentar detectar hojas automáticamente ──
    sheet_map = {s.lower(): s for s in wb.sheetnames}

    def get_sheet(*candidates):
        for c in candidates:
            if c.lower() in sheet_map:
                return wb[sheet_map[c.lower()]]
        # fallback: primera hoja
        return wb.active

    ws_vcp     = get_sheet("vcp", "cuotapartes", "valores")
    ws_cartera = get_sheet("cartera", "portafolio", "inversiones")
    ws_futuros = get_sheet("futuros", "rofex", "derivados")
    ws_cpd     = get_sheet("cpd", "cheques", "vencimientos")
    ws_atrib   = get_sheet("atribucion", "rendimiento", "variacion")

    # ─────────────────────────────────────────────────────────────
    # VCP por clase  ← AJUSTÁ estas celdas
    # ─────────────────────────────────────────────────────────────
    # Ejemplo: ws_vcp.cell(row=5, column=3).value
    vcp_data = {
        "A": {
            "vcp":       safe_float(cell(ws_vcp, 5,  3)),   # ← C5 en hoja VCP
            "vcp_prev":  safe_float(cell(ws_vcp, 5,  4)),   # ← D5
            "vcp_abril": safe_float(cell(ws_vcp, 5,  5)),   # ← E5
            "vcp_dic":   safe_float(cell(ws_vcp, 5,  6)),   # ← F5
            "saldo_cp":  safe_float(cell(ws_vcp, 5,  7)),   # ← G5
            "saldo_cp_prev": safe_float(cell(ws_vcp, 5, 8)),
            "ingresos_hoy":  safe_float(cell(ws_vcp, 5, 9)),
            "egresos_hoy":   safe_float(cell(ws_vcp, 5, 10)),
        },
        "B": {
            "vcp":       safe_float(cell(ws_vcp, 6,  3)),
            "vcp_prev":  safe_float(cell(ws_vcp, 6,  4)),
            "vcp_abril": safe_float(cell(ws_vcp, 6,  5)),
            "vcp_dic":   safe_float(cell(ws_vcp, 6,  6)),
            "saldo_cp":  safe_float(cell(ws_vcp, 6,  7)),
            "saldo_cp_prev": safe_float(cell(ws_vcp, 6, 8)),
            "ingresos_hoy":  safe_float(cell(ws_vcp, 6, 9)),
            "egresos_hoy":   safe_float(cell(ws_vcp, 6, 10)),
        },
        "C": {
            "vcp":       safe_float(cell(ws_vcp, 7,  3)),
            "vcp_prev":  safe_float(cell(ws_vcp, 7,  4)),
            "vcp_abril": safe_float(cell(ws_vcp, 7,  5)),
            "vcp_dic":   safe_float(cell(ws_vcp, 7,  6)),
            "saldo_cp":  safe_float(cell(ws_vcp, 7,  7)),
            "saldo_cp_prev": safe_float(cell(ws_vcp, 7, 8)),
            "ingresos_hoy":  safe_float(cell(ws_vcp, 7, 9)),
            "egresos_hoy":   safe_float(cell(ws_vcp, 7, 10)),
        },
    }

    # ─────────────────────────────────────────────────────────────
    # Atribución VCP  ← AJUSTÁ estas celdas
    # ─────────────────────────────────────────────────────────────
    atrib_items = []
    # Asume filas 2..12 con columnas: nombre, tipo, bps, delta_pesos, detalle
    for r in range(2, 14):
        nombre = cell(ws_atrib, r, 1)
        if not nombre:
            continue
        atrib_items.append({
            "n":   str(nombre),
            "t":   str(cell(ws_atrib, r, 2) or ""),
            "bps": safe_float(cell(ws_atrib, r, 3)) or 0,
            "dp":  safe_float(cell(ws_atrib, r, 4)),
            "det": str(cell(ws_atrib, r, 5) or ""),
        })

    delta_total = safe_float(cell(ws_atrib, 14, 3)) or sum(i["bps"] for i in atrib_items)

    # ─────────────────────────────────────────────────────────────
    # Cartera de inversiones  ← AJUSTÁ estas celdas
    # ─────────────────────────────────────────────────────────────
    cartera = {
        "cpd_pesos":       safe_float(cell(ws_cartera, 3, 3)),
        "cpd_pesos_prev":  safe_float(cell(ws_cartera, 3, 4)),
        "cpd_usd":         safe_float(cell(ws_cartera, 4, 3)),
        "cpd_usd_prev":    safe_float(cell(ws_cartera, 4, 4)),
        "pases_cauciones": safe_float(cell(ws_cartera, 5, 3)),
        "pases_cauciones_prev": safe_float(cell(ws_cartera, 5, 4)),
        "money_market":    safe_float(cell(ws_cartera, 6, 3)),
        "money_market_prev": safe_float(cell(ws_cartera, 6, 4)),
    }

    # ─────────────────────────────────────────────────────────────
    # Vencimientos CPD (próximos 10 días)  ← AJUSTÁ estas celdas
    # ─────────────────────────────────────────────────────────────
    cpd_venc = []
    DIAS_SEMANA = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    for r in range(2, 13):
        fecha_val = cell(ws_cpd, r, 1)
        monto_val = safe_float(cell(ws_cpd, r, 2))
        if not fecha_val or not monto_val:
            continue
        # Si la fecha viene como datetime de Excel
        if isinstance(fecha_val, datetime):
            dia_semana = DIAS_SEMANA[fecha_val.weekday()]
            fecha_str  = fecha_val.strftime("%d/%m")
        else:
            fecha_str  = str(fecha_val)
            dia_semana = ""
        cpd_venc.append({"fecha": fecha_str, "dia": dia_semana, "monto": monto_val})

    # ─────────────────────────────────────────────────────────────
    # Futuros ROFEX  ← AJUSTÁ estas celdas
    # ─────────────────────────────────────────────────────────────
    futuros = []
    for r in range(2, 8):
        contrato = cell(ws_futuros, r, 1)
        if not contrato:
            continue
        futuros.append({
            "contrato":   str(contrato),
            "cantidad":   safe_float(cell(ws_futuros, r, 2)),
            "cierre":     safe_float(cell(ws_futuros, r, 3)),
            "total_hoy":  safe_float(cell(ws_futuros, r, 4)),
            "total_prev": safe_float(cell(ws_futuros, r, 5)),
        })

    return {
        "cuotapartes": vcp_data,
        "atribucion": {
            "delta_total": delta_total,
            "items": atrib_items,
        },
        "cartera": cartera,
        "cpd_vencimientos": cpd_venc,
        "futuros": futuros,
    }


def parse_balance(path: Path) -> dict:
    """
    Lee el archivo Balance_DD-MM-YYYY.xls.

    ┌─────────────────────────────────────────────────────────────┐
    │  CONFIGURACIÓN — ajustá las referencias de celdas aquí     │
    └─────────────────────────────────────────────────────────────┘
    """
    # xlrd para .xls (formato antiguo)
    ext = path.suffix.lower()
    if ext == ".xls":
        book = xlrd.open_workbook(str(path))
        print(f"  Hojas en {path.name}: {book.sheet_names()}")

        def xls_cell(sheet_name, row, col):
            try:
                sh = book.sheet_by_name(sheet_name)
            except xlrd.XLRDError:
                sh = book.sheet_by_index(0)
            try:
                v = sh.cell_value(row, col)
                return v if v != '' else None
            except Exception:
                return None

        # ── AJUSTÁ estas referencias (base 0: fila 0 = fila 1 en Excel) ──
        activo_total      = safe_float(xls_cell("Balance", 4, 2))   # C5
        activo_total_prev = safe_float(xls_cell("Balance", 4, 3))   # D5
        pasivo_total      = safe_float(xls_cell("Balance", 14, 2))
        pasivo_total_prev = safe_float(xls_cell("Balance", 14, 3))
        pn_hoy            = safe_float(xls_cell("Balance", 16, 2))
        pn_prev           = safe_float(xls_cell("Balance", 16, 3))
        creditos_hoy      = safe_float(xls_cell("Balance", 8, 2))
        creditos_prev     = safe_float(xls_cell("Balance", 8, 3))
        fecha_actual_raw  = xls_cell("Balance", 0, 2)
        fecha_prev_raw    = xls_cell("Balance", 0, 3)
        duration          = safe_float(xls_cell("Balance", 2, 2))

    else:  # .xlsx
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"  Hojas en {path.name}: {wb.sheetnames}")
        ws = wb.active

        # ── AJUSTÁ estas referencias ──
        activo_total      = safe_float(cell(ws, 5,  3))
        activo_total_prev = safe_float(cell(ws, 5,  4))
        pasivo_total      = safe_float(cell(ws, 15, 3))
        pasivo_total_prev = safe_float(cell(ws, 15, 4))
        pn_hoy            = safe_float(cell(ws, 17, 3))
        pn_prev           = safe_float(cell(ws, 17, 4))
        creditos_hoy      = safe_float(cell(ws, 9,  3))
        creditos_prev     = safe_float(cell(ws, 9,  4))
        fecha_actual_raw  = cell(ws, 1, 3)
        fecha_prev_raw    = cell(ws, 1, 4)
        duration          = safe_float(cell(ws, 3,  3))

    lc = None
    if activo_total and pasivo_total and pasivo_total != 0:
        lc = round(activo_total / pasivo_total, 1)

    def fmt_fecha(v):
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        if v:
            return str(v)
        return "—"

    return {
        "meta": {
            "fecha_actual":   fmt_fecha(fecha_actual_raw),
            "fecha_anterior": fmt_fecha(fecha_prev_raw),
            "duration_dias":  duration,
            "participes":     4,            # ajustá o leelo del xlsx
        },
        "patrimonial": {
            "activo_total":       activo_total,
            "activo_total_prev":  activo_total_prev,
            "pasivo_total":       pasivo_total,
            "pasivo_total_prev":  pasivo_total_prev,
            "patrimonio_neto":    pn_hoy,
            "patrimonio_neto_prev": pn_prev,
            "creditos_totales":   creditos_hoy,
            "creditos_totales_prev": creditos_prev,
            "liquidez_corriente": lc,
        },
    }


# ─────────────────────────────────────────────
# MAIN — empareja archivos por fecha y genera JSON
# ─────────────────────────────────────────────
def main():
    # Recopila todos los archivos en raw/
    gestiones = {}
    balances  = {}

    for f in sorted(RAW_DIR.glob("*")):
        if not f.is_file():
            continue
        date = date_from_filename(f.stem)
        if not date:
            print(f"  ⚠️  No pude extraer fecha de '{f.name}' — omitido")
            continue

        name_lower = f.name.lower()
        if "gestion" in name_lower or "momentum" in name_lower:
            gestiones[date] = f
        elif "balance" in name_lower:
            balances[date] = f
        else:
            print(f"  ⚠️  Archivo no reconocido: '{f.name}' — omitido")

    all_dates = sorted(set(gestiones) | set(balances))
    if not all_dates:
        print("No se encontraron archivos en raw/")
        return

    generated = []

    for date in all_dates:
        print(f"\n📅 Procesando {date}")
        result = {}

        if date in balances:
            print(f"  Balance: {balances[date].name}")
            try:
                result.update(parse_balance(balances[date]))
            except Exception as e:
                print(f"  ❌ Error en balance: {e}")
        else:
            print(f"  ⚠️  Sin archivo de balance para {date}")

        if date in gestiones:
            print(f"  Gestión: {gestiones[date].name}")
            try:
                result.update(parse_gestion(gestiones[date]))
            except Exception as e:
                print(f"  ❌ Error en gestión: {e}")
        else:
            print(f"  ⚠️  Sin archivo de gestión para {date}")

        out_path = DATA_DIR / f"{date}.json"
        with open(out_path, "w", encoding="utf-8") as fh:
            json.dump(result, fh, ensure_ascii=False, indent=2)
        print(f"  ✅ Guardado: data/{date}.json")
        generated.append(date)

    # Actualiza index.json
    index = {
        "dates":  generated,
        "latest": generated[-1] if generated else None,
    }
    with open(DATA_DIR / "index.json", "w", encoding="utf-8") as fh:
        json.dump(index, fh, ensure_ascii=False, indent=2)
    print(f"\n✅ index.json actualizado — {len(generated)} días disponibles")


if __name__ == "__main__":
    main()
